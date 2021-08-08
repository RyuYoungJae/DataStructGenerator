import openpyxl as xl
import glob2
import Util

class DataFileRead(object):
    def Read(self, path):
        result = {}
        xlFiles = glob2.glob(path)
        for file in xlFiles:
            wb = xl.load_workbook(file)
            for sheetName in wb.sheetnames:
                if sheetName != "Define":
                    continue

                factors = {}
                sheet = wb[sheetName]
                for row in sheet.iter_rows(min_row=1):
                    variable = []
                    for cell in row:
                        if cell.value == "Column" or cell.value == "" : 
                            break
                        variable.append(cell.value)
                    if len(variable) < 2 or len(variable) > 2:
                        continue
                    factors[variable[0]] = variable[1]

                util = Util.Util()
                fileName = util.GetFileName(file)
                result[fileName[0]] = factors
            wb.close()
        return result


