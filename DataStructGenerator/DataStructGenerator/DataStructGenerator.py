import openpyxl as xl
import glob2

path = "../../Data/*.xlsx"
xlFiles = glob2.glob(path)

for file in xlFiles:
    wb = xl.load_workbook(file)
    for sheetName in wb.sheetnames:
        if sheetName != "Define":
            continue

        sheet = wb[sheetName]
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                print("[", cell.value, "]")

    wb.close()